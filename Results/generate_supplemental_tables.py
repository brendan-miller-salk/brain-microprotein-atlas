#!/usr/bin/env python3
"""
Generate Master Supplementary Tables         "S14": "Outer Mitochondrial Membrane Differential Proteins - Proteomics analysis of mitochondrial membrane-associated microproteins",
        "S15": "TMT ROSMAP Donor Metadata - Demographic characteristics of study participants by sex and diagnostic group",
        "S16": "psORF Actin-Related Microproteins - Pseudogene-derived open reading frames encoding actin-related microprotein sequences",
        "S17": "RNA-Binding Protein Splice-Site Motifs - Per-junction RBP motif enrichment at smORF-spanning splice sites"
    }
    return titles.get(table_id, "Supplementary Data") File
==============================================================================
Purpose:
    - Consolidate all supplementary data tables into a single Excel workbook
    - Each CSV file becomes a separate tab (S1, S2, etc.)
    - Provides error handling for missing files
    - Creates formatted Excel output with proper sheet names

Author: Generated for AD Microprotein Analysis
Date: September 2025
==============================================================================
"""

import pandas as pd
import os
import argparse
from pathlib import Path
import sys
import shutil
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

def get_brief_description(table_id, full_description):
    """Create brief descriptions for Excel tab names (Excel has 31 character limit)"""
    brief_descriptions = {
        "S1": "Discovery",
        "S2": "Proteomics", 
        "S3": "Tryptic Peptides",
        "S4": "RP3 Results",
        "S5": "ShortStop",
        "S6": "scRNA Enrich",
        "S7": "Short-Read RNA",
        "S8": "TSS Motifs",
        "S9": "RBP Splice Sites",
        "S10": "Coexp AD T1",
        "S11": "Coexp nonAD T1",
        "S12": "Coexp AD T2", 
        "S13": "Coexp Comb T2",
        "S14": "OMM Diff",
        "S15": "Long-Read RNA",
        "S16": "RP3 psORFs",
        "S17": "TMT Donor Meta",
        "S18": "smORF Type Defs"
    }
    return brief_descriptions.get(table_id, "Data")

def get_full_table_title(table_id):
    """Get full descriptive titles for each supplementary table"""
    titles = {
        "S1": "Brain Microproteins Discovery Summary - Comprehensive annotation and classification of microproteins identified across all analyses",
        "S2": "Proteomics Results Summary - Mass spectrometry evidence for microprotein expression with peptide-spectrum matches",
        "S3": "Tryptic Peptides Summary - Unique tryptic peptide sequences identified in microprotein mass spectrometry analysis", 
        "S4": "RP3 Results Summary - Ribosome footprint quantification evidence for microprotein translation with proteomics validation",
        "S5": "ShortStop Microproteins Summary - Machine learning classification and annotation of short open reading frames",
        "S6": "Single-cell RNA Enrichment Summary - Cell-type specific expression patterns of microproteins in brain tissue",
        "S7": "Short-Read RNA-seq Transcriptomics Results - Differential expression analysis and main-ORF/smORF co-expression statistics (ROSMAP DLPFC)",
        "S8": "Transcription Start Site Motif Prevalence - Promoter motif prevalence at predicted smORF transcription start sites",
        "S9": "RNA-Binding Protein Splice-Site Motifs - Per-junction RBP motif enrichment at smORF-spanning splice sites",
        "S10": "Co-expression Analysis AD (Target uORF-1-MKKS) - Gene co-expression networks in Alzheimer's Disease for uORF-1-MKKS",
        "S11": "Co-expression Analysis Non-AD (Target uORF-1-MKKS) - Gene co-expression networks in control samples for uORF-1-MKKS",
        "S12": "Co-expression Analysis AD (Target MKKS-Reviewed-570aa) - Gene co-expression networks in Alzheimer's Disease for reviewed MKKS-570aa", 
        "S13": "Co-expression Analysis Non-AD (Target MKKS-Reviewed-570aa) - Non-AD gene co-expression networks for reviewed MKKS-570aa",
        "S14": "Outer Mitochondrial Membrane Differential Proteins - Proteomics analysis of mitochondrial membrane-associated microproteins",
        "S15": "Long-Read RNA-seq Transcriptomics Results - Long-read sequencing analysis of microprotein transcript expression",
        "S16": "RP3 Pseudogene-derived ORFs - Ribosome profiling evidence for translation of pseudogene-derived microproteins",
        "S17": "TMT ROSMAP Donor Metadata - Demographic characteristics of study participants by sex and diagnostic group",
        "S18": "smORF Type Definitions - Comprehensive definitions of small open reading frame (smORF) classification categories"
    }
    return titles.get(table_id, f"Supplementary Table {table_id}")

def add_status_column(df):
    """Add Status column as first column based on Database classification"""
    try:
        # Look for Database column (case insensitive)
        database_col = None
        for col in df.columns:
            if col.lower() in ['database', 'db']:
                database_col = col
                break
        
        if database_col is not None:
            # Create Status column
            def classify_status(database_value):
                if pd.isna(database_value):
                    return "Unknown"
                database_str = str(database_value).strip()
                if 'Swiss-Prot' in database_str:
                    return "Reviewed"
                elif any(term in database_str for term in ['Salk', 'TrEMBL', 'Unreviewed']):
                    return "Unreviewed"
                else:
                    return "Unknown"
            
            # Apply classification
            status_column = df[database_col].apply(classify_status)
            
            # Insert Status as first column
            df.insert(0, 'Status', status_column)
            print(f"  🏷️  Added Status column based on {database_col}")
            
            # Show distribution
            status_counts = status_column.value_counts()
            for status, count in status_counts.items():
                print(f"    {status}: {count:,} entries")
        else:
            print(f"  ⚠️  No Database column found - Status column not added")
    
    except Exception as e:
        print(f"  ⚠️  Warning: Could not add Status column - {str(e)}")
    
    return df

def move_click_ucsc_to_front(df):
    """Move CLICK_UCSC column to the front if it exists"""
    try:
        # Look for CLICK_UCSC column (case insensitive)
        click_ucsc_col = None
        for col in df.columns:
            if col.upper() == 'CLICK_UCSC':
                click_ucsc_col = col
                break
        
        if click_ucsc_col is not None:
            # Remove the column and insert it at the beginning
            ucsc_data = df[click_ucsc_col]
            df = df.drop(click_ucsc_col, axis=1)
            df.insert(0, click_ucsc_col, ucsc_data)
            print(f"  🔗 Moved {click_ucsc_col} column to front")
        
    except Exception as e:
        print(f"  ⚠️  Warning: Could not move CLICK_UCSC column - {str(e)}")
    
    return df

def filter_s8_nanopore_data(df, table_id):
    """Remove rows where nanopore_baseMean is missing or blank for S8 table"""
    try:
        if table_id == "S15":  # Long-Read RNA (shifted to S15)
            if 'nanopore_baseMean' in df.columns:
                before_count = len(df)
                # Remove rows where nanopore_baseMean is NaN, None, or empty string
                df = df.dropna(subset=['nanopore_baseMean'])
                df = df[df['nanopore_baseMean'] != '']
                df = df[df['nanopore_baseMean'] != 0]  # Also remove zeros if needed
                after_count = len(df)
                removed_count = before_count - after_count
                if removed_count > 0:
                    print(f"  🧹 Removed {removed_count:,} rows with missing/blank nanopore_baseMean")
                else:
                    print(f"  ✅ All rows have valid nanopore_baseMean values")
            else:
                print(f"  ⚠️  nanopore_baseMean column not found in S8 table")
    
    except Exception as e:
        print(f"  ⚠️  Warning: Could not filter S8 nanopore data - {str(e)}")
    
    return df

def apply_table_sorting(df, table_id):
    """Apply table-specific sorting based on requirements"""
    try:
        if table_id == "S2":  # Proteomics
            if 'TMT_qvalue_50pct_missing' in df.columns:
                df = df.sort_values('TMT_qvalue_50pct_missing', ascending=True)
                print(f"  📊 Sorted by TMT_qvalue_50pct_missing (lowest to highest)")
        
        elif table_id == "S4":  # RP3
            if 'RP3_MM_Amb' in df.columns:
                df = df.sort_values('RP3_MM_Amb', ascending=False)
                print(f"  📊 Sorted by RP3_MM_Amb (highest to lowest)")
        
        elif table_id == "S5":  # ShortStop
            if 'ShortStop Score' in df.columns:
                df = df.sort_values('ShortStop Score', ascending=False)
                print(f"  📊 Sorted by ShortStop Score (highest to lowest)")
        
        elif table_id == "S7":  # Short-Read RNA
            if 'rosmapRNA_padj' in df.columns:
                df = df.sort_values('rosmapRNA_padj', ascending=True)
                print(f"  📊 Sorted by rosmapRNA_padj (lowest to highest)")
    
    except Exception as e:
        print(f"  ⚠️  Warning: Could not apply sorting - {str(e)}")
    
    return df

def create_column_definitions_worksheet(writer):
    """Create Column Definitions worksheet with detailed column descriptions"""
    
    # Define all column definitions
    column_definitions = [
        ("S1", "Annotation Status", "Whether the microprotein was found in MS or RiboSeq/ShortStop"),
        ("S1", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S1", "Microprotein Length", "Length of the encoded microprotein (amino acids)."),
        ("S1", "Microprotein Sequence", "Amino acid sequence of the microprotein (smORF translation product)."),
        ("S1", "MS Evidence Type", "DDA or DIA (or both)"),
        ("S1", "DDA Grade", "Best PROSIT confidence tier (Strong > Moderate > Weak > Insufficient) across tryptic peptides for DDA evidence quality."),
        ("S1", "Parent Gene", "Gene that contains or is associated with the smORF."),
        ("S1", "smORF Class", "Category of smORF (e.g., uORF, dORF, intergenic, lncRNA-derived)."),
        ("S1", "smORF Coordinates", "Genomic coordinates of the smORF (chr:start-end:strand)."),
        ("S10", "abs_corr", "Absolute value of the correlation coefficient."),
        ("S10", "correlation", "Correlation coefficient with a reference variable. R"),
        ("S10", "enst_clean", "Cleaned version of Ensembl transcript ID for analysis."),
        ("S10", "enst_id", "Ensembl transcript identifier."),
        ("S10", "external_gene_name", "Gene name associated with the transcript."),
        ("S11", "abs_corr", "Absolute value of the correlation coefficient."),
        ("S11", "correlation", "Correlation coefficient with a reference variable. R"),
        ("S11", "enst_clean", "Cleaned version of Ensembl transcript ID for analysis."),
        ("S11", "enst_id", "Ensembl transcript identifier."),
        ("S11", "external_gene_name", "Gene name associated with the transcript."),
        ("S12", "abs_corr", "Absolute value of the correlation coefficient."),
        ("S12", "correlation", "Correlation coefficient with a reference variable. R"),
        ("S12", "enst_clean", "Cleaned version of Ensembl transcript ID for analysis."),
        ("S12", "enst_id", "Ensembl transcript identifier."),
        ("S12", "external_gene_name", "Gene name associated with the transcript."),
        ("S13", "is_mkks", "Boolean flag indicating whether the protein is Micro-MKKS."),
        ("S13", "label", "Label for visualization or annotation purposes."),
        ("S13", "log2FC", "Log2 fold-change of protein abundance between groups."),
        ("S13", "negLog10P", "Negative log10-transformed p-value for visualization."),
        ("S13", "padj", "Adjusted p-value using FDR or similar correction."),
        ("S13", "ProteinName", "Name of the protein or microprotein being analyzed."),
        ("S13", "pval", "Raw p-value for the statistical test."),
        ("S13", "significant", "Indicates whether the protein meets significance thresholds."),
        ("S3", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S3", "Database", "Database annotation source (e.g., Salk, Swiss-Prot, TrEMBL); Salk is an arbitrary name from custom GTFtoFASTA in silico translation tool"),
        ("S3", "gene_name", "Name of the parent gene name of the UNREVIEWED ORF gene; reviewed entries will be blank --> see gene_symbol"),
        ("S3", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs"),
        ("S3", "ROSMAP_BulkRNAseq_CPM", "Counts-per-million normalized RNA abundance."),
        ("S3", "rosmapRNA_baseMean", "Mean normalized RNA expression across samples."),
        ("S3", "rosmapRNA_body", "Whether the entire gene body is differentialy expressed or just the unique CDS of the gene body"),
        ("S3", "rosmapRNA_log2FoldChange", "Log2 fold-change between disease vs. control."),
        ("S3", "rosmapRNA_non_smorf_hit", "Indicator of non-smORF overlapping signal."),
        ("S3", "rosmapRNA_padj", "Benjamini–Hochberg adjusted p-value (FDR)."),
        ("S3", "rosmapRNA_pvalue", "Raw p-value for differential expression."),
        ("S3", "rosmapRNA_stat", "Wald test statistic for DESeq2 analysis."),
        ("S3", "sequence", "Amino acid sequence of the ORF product."),
        ("S3", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)"),
        ("S2", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S2", "Database", "Database annotation source (e.g., Salk, Swiss-Prot, TrEMBL); Salk is an arbitrary name from custom GTFtoFASTA in silico translation tool"),
        ("S2", "gene_name", "Name of the parent gene name of the UNREVIEWED ORF gene; reviewed entries will be blank --> see gene_symbol"),
        ("S2", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs"),
        ("S2", "protein_class_length", "Length of the protein class (Microprotein or Lon protein greater than 150 aa)."),
        ("S2", "protein_length", "Protein length in amino acids."),
        ("S2", "sequence", "Amino acid sequence of the ORF product."),
        ("S2", "smorf_type", "Type of smORF (uORF, dORF, intergenic, etc.)."),
        ("S2", "start_codon", "ATG or nonATG"),
        ("S2", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)"),
        ("S2", "TMT_log2fc_50pct_missing", "Log2 fold-change of TMT intensity (AD vs. Control) — 50% missing-value threshold (more inclusive)."),
        ("S2", "TMT_pvalue_50pct_missing", "Raw p-value for the TMT proteomics comparison — 50% missing-value threshold."),
        ("S2", "TMT_qvalue_50pct_missing", "BH-adjusted q-value for the TMT proteomics comparison — 50% missing-value threshold."),
        ("S2", "TMT_log2fc_0pct_missing", "Log2 fold-change of TMT intensity (AD vs. Control) — 0% missing-value threshold (stringent; no imputation)."),
        ("S2", "TMT_pvalue_0pct_missing", "Raw p-value for the TMT proteomics comparison — 0% missing-value threshold."),
        ("S2", "TMT_qvalue_0pct_missing", "BH-adjusted q-value for the TMT proteomics comparison — 0% missing-value threshold."),
        ("S2", "rate_control", "MS detection rate in Control donors (fraction of Control samples where the protein was detected)."),
        ("S2", "rate_ad", "MS detection rate in AD donors (fraction of AD samples where the protein was detected)."),
        ("S2", "total_razor_spectral_counts", "Total razor peptide spectral counts assigned."),
        ("S2", "total_unique_spectral_counts", "Total unique peptide spectral counts assigned."),
        ("S3", "SA_degrees", "Spectral angle in degrees from PROSIT for each tryptic peptide (bracketed list)."),
        ("S3", "SA_normalized", "Normalized spectral angle from PROSIT for each tryptic peptide (bracketed list)."),
        ("S7", "correlation_mainORF_nonAD_rosmap", "Pearson correlation between smORF and main-ORF transcript expression in non-AD donors (ROSMAP, batch-corrected VST)."),
        ("S7", "correlation_mainORF_AD_rosmap", "Pearson correlation between smORF and main-ORF transcript expression in AD donors (ROSMAP, batch-corrected VST)."),
        ("S7", "rosmap_lrt_additive_p", "LRT p-value for the additive model (smORF ~ main + group): tests whether AD diagnosis independently predicts smORF expression."),
        ("S7", "rosmap_lrt_interaction_p", "LRT p-value for the interaction model (smORF ~ main * group): tests whether the main-ORF/smORF relationship differs by diagnosis."),
        ("S3", "Match_coverage", "Ion match coverage from MS/MS fragmentation for each tryptic peptide (bracketed list)."),
        ("S3", "Match_coverage_pct", "Ion match coverage percentage from MS/MS fragmentation for each tryptic peptide (bracketed list)."),
        ("S3", "Confidence", "4-tiered confidence classification based on spectral angle and match coverage combination."),
        ("S7", "sequence", "Full amino acid sequence of the microprotein."),
        ("S7", "protein_id", "Unique protein identifier from the mass spectrometry database."),
        ("S7", "peptide_sequence", "List of tryptic peptide sequences identified for this protein."),
        ("S7", "start", "List of start positions for each tryptic peptide within the protein sequence."),
        ("S7", "end", "List of end positions for each tryptic peptide within the protein sequence."),
        ("S7", "is_unique_peptide", "Boolean indicating whether the peptide(s) contain at least one unique to this protein."),
        ("S5", "Annotation Status", "Annotated vs. novel classification status."),
        ("S5", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S5", "Gene Body (Name)", "Name of the parent gene body for unreviewed ORFs; also referred to as 'gene_name' in other S Tables."),
        ("S5", "Microprotein Sequence", "Predicted amino acid sequence."),
        ("S5", "ShortStop Label", "Label assigned by ShortStop ML classifier."),
        ("S5", "ShortStop Score", "Confidence score from the ShortStop classifier."),
        ("S5", "smORF Class", "Category of smORF (e.g., uORF, dORF, intergenic, lncRNA-derived)."),
        ("S5", "smORF ID", "Unique smORF identifier."),
        ("S4", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S4", "Database", "Database annotation source (e.g., Salk, Swiss-Prot, TrEMBL); Salk is an arbitrary name from custom GTFtoFASTA in silico translation tool"),
        ("S4", "gene_name", "Name of the parent gene name of the UNREVIEWED ORF gene; reviewed entries will be blank --> see gene_symbol"),
        ("S4", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs"),
        ("S4", "protein_length", "Protein length in amino acids."),
        ("S4", "RiboCode", "RiboCode-based ORF translation call."),
        ("S4", "RP3_Amb", "Ambiguous counts from RP3"),
        ("S4", "RP3_Default", "Non ambiguous nor multi-mapping counts from RP3"),
        ("S4", "RP3_MM", "Multi-mapping counts from RP3"),
        ("S4", "RP3_MM_Amb", "Both ambiguous nor multi-mapping counts from RP3"),
        ("S4", "sequence", "Amino acid sequence of the ORF product."),
        ("S4", "smorf_type", "Type of smORF (uORF, dORF, intergenic, etc.)."),

        ("S4", "cell_type_general", "scRNA-Seq cell type annotations from PMID: 37774677"),
        ("S4", "celltype", "scRNA-Seq cell subtype annotations from PMID: 37774677"),
        ("S4", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S4", "cluster_id", "Cluster ID from PMID: 37774677"),
        ("S4", "coef", "Coefficient to pathology by PMID: 37774677"),
        ("S4", "Database", "Database annotation source (e.g., Salk, Swiss-Prot, TrEMBL); Salk is an arbitrary name from custom GTFtoFASTA in silico translation tool"),
        ("S4", "F", "F score from PMID: 37774677"),
        ("S4", "gene", "Gene body expressed by cell type from PMID: 37774677"),
        ("S4", "gene_id", "Unique CDS identifier of protein sequence corresponding to gene body"),
        ("S4", "logCPM", "logCPM of gene expression counts from PMID: 37774677"),
        ("S4", "logFC", "fold change of gene expression counts from PMID: 37774677"),
        ("S4", "p_adj.glb", "adjusted p value provided by PMID: 37774677"),
        ("S4", "p_adj.loc", "adjusted p value provided by PMID: 37774677"),
        ("S4", "p_val", "raw p value provided by PMID: 37774677"),
        ("S4", "sequence", "Multi-mapping counts from RP3"),
        ("S4", "smorf_type", "Type of smORF (uORF, dORF, intergenic, etc.)."),
        ("S4", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)"),
        ("S4", "total_unique_spectral_counts", "Total unique peptide spectral counts assigned."),
        ("S6", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S6", "Database", "Database annotation source (e.g., Salk, Swiss-Prot, TrEMBL); Salk is an arbitrary name from custom GTFtoFASTA in silico translation tool"),
        ("S6", "gene_name", "Name of the parent gene name of the UNREVIEWED ORF gene; reviewed entries will be blank --> see gene_symbol"),
        ("S6", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs"),
        ("S6", "protein_length", "Protein length in amino acids."),
        ("S6", "RiboCode", "RiboCode-based ORF translation call."),
        ("S6", "RP3_Amb", "Ambiguous counts from RP3"),
        ("S6", "RP3_Default", "Non ambiguous nor multi-mapping counts from RP3"),
        ("S6", "RP3_MM", "Multi-mapping counts from RP3"),
        ("S6", "RP3_MM_Amb", "Both ambiguous nor multi-mapping counts from RP3"),
        ("S6", "sequence", "Multi-mapping counts from RP3"),
        ("S6", "smorf_type", "Type of smORF (uORF, dORF, intergenic, etc.)."),
        ("S6", "start_codon", "ATG or nonATG"),
        ("S6", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)"),
        ("S6", "total_razor_spectral_counts", "Total razor peptide spectral counts assigned."),
        ("S6", "total_unique_spectral_counts", "Total unique peptide spectral counts assigned."),
        ("S15", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S15", "Database", "Database annotation source (e.g., Salk, Swiss-Prot, TrEMBL); Salk is an arbitrary name from custom GTFtoFASTA in silico translation tool"),
        ("S15", "gene_name", "Name of the parent gene name of the UNREVIEWED ORF gene; reviewed entries will be blank --> see gene_symbol"),
        ("S15", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs"),
        ("S15", "nanopore_baseMean", "Mean normalized expression from Nanopore long-read RNA-seq."),
        ("S15", "nanopore_log2FoldChange", "Log2 fold-change between disease vs. control from Nanopore data."),
        ("S15", "nanopore_padj", "Adjusted p-value (FDR) for Nanopore data."),
        ("S15", "nanopore_pvalue", "Raw p-value for Nanopore-based differential expression."),
        ("S15", "protein_length", "Protein length in amino acids."),
        ("S15", "sequence", "Amino acid sequence of the ORF product."),
        ("S15", "smorf_type", "Type of smORF (uORF, dORF, intergenic, etc.)."),
        ("S15", "start_codon", "ATG or nonATG"),
        ("S15", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)"),
        ("S15", "total_razor_spectral_counts", "Total razor peptide spectral counts assigned."),
        ("S15", "total_unique_spectral_counts", "Total unique peptide spectral counts assigned."),
        ("S9", "abs_corr", "Absolute value of the correlation coefficient."),
        ("S9", "correlation", "Correlation coefficient with a reference variable. R"),
        ("S9", "enst_clean", "Cleaned version of Ensembl transcript ID for analysis."),
        ("S9", "enst_id", "Ensembl transcript identifier."),
        ("S9", "external_gene_name", "Gene name associated with the transcript."),
        ("S14", "ProteinName", "Name of the outer mitochondrial membrane protein."),
        ("S14", "log2FC", "Log2 fold-change of protein abundance between AD vs Control groups."),
        ("S14", "pval", "Raw p-value for the differential abundance test."),
        ("S14", "padj", "FDR-adjusted p-value for multiple testing correction."),
        ("S14", "negLog10P", "Negative log10-transformed p-value for volcano plot visualization."),
        ("S14", "significant", "Boolean indicating whether the protein meets significance thresholds (padj < 0.05 and |log2FC| > 1)."),
        ("S14", "is_mkks", "Boolean flag indicating whether the protein is related to MKKS/Bardet-Biedl syndrome pathway."),
        ("S14", "label", "Annotation label for visualization and grouping purposes."),
        ("S17", "Biological Sex", "Biological sex of study participants (Male/Female)."),
        ("S17", "Diagnosis", "Diagnostic classification (AD = Alzheimer's Disease, AsymAD = Asymptomatic Alzheimer's Disease, Control = Control)."),
        ("S17", "n", "Number of participants in each demographic group."),
        ("S17", "Mean Age", "Mean age of participants in years."),
        ("S17", "Standard Deviation Age", "Standard deviation of age within each group."),
        ("S17", "Median Age", "Median age of participants in years."),
        ("S18", "General smORF Type", "Broad classification grouping of the smORF (e.g., Upstream ORF, Downstream ORF, Internal ORF, lncRNA, Pseudogene ORF, Alternative Frame ORF)."),
        ("S18", "smORF Type", "Specific smORF subtype abbreviation (e.g., uORF, dORF, iORF, lncRNA, psORF, AltORF / TrEMBL)."),
        ("S18", "Definition", "Comprehensive definition and description of the precise smORF type classification."),
        ("S9", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)."),
        ("S9", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S9", "sequence", "Amino acid sequence of the ORF product."),
        ("S9", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs."),
        ("S9", "gene_id", "Ensembl gene identifier for the parent gene."),
        ("S9", "junction", "Splice junction identifier (e.g., junction_1_acceptor or junction_1_donor)."),
        ("S9", "rbp_motifs", "List of RNA-binding protein motifs enriched at this splice junction."),
        ("S8", "Status", "Whether the sequence is Reviewed (Swiss-Prot) or Unreviewed (TrEMBL and Novel; Salk-generated database or TrEMBL)."),
        ("S8", "CLICK_UCSC", "Direct link to the genomic region in the UCSC Genome Browser for visualization."),
        ("S8", "sequence", "Amino acid sequence of the ORF product."),
        ("S8", "gene_symbol", "Standardized gene symbol for reviewed and unreviewed ORFs."),
        ("S8", "promoter_motif", "Transcription factor binding motif identified at the predicted transcription start site.")
    ]
    
    # Create DataFrame
    col_def_df = pd.DataFrame(column_definitions, columns=['Table', 'Column Name', 'Description'])
    col_def_df.to_excel(writer, sheet_name='Column Definitions', index=False)
    
    # Style the Column Definitions worksheet
    col_def_worksheet = writer.sheets['Column Definitions']
    
    # Add title
    col_def_worksheet.insert_rows(1)
    col_def_worksheet.insert_rows(1)
    col_def_worksheet.insert_rows(1)  # Three rows for title
    
    title_cell = col_def_worksheet.cell(row=1, column=1)
    title_cell.value = "COLUMN DEFINITIONS - Brain Microproteins Supplementary Data"
    
    # Style title
    title_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")  # Deep purple
    title_font = Font(color="FFFFFF", bold=True, size=16)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    col_def_worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    col_def_worksheet.row_dimensions[1].height = 50
    
    # Style headers (row 4)
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")  # Purple
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, 4):  # 3 columns
        cell = col_def_worksheet.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Style data rows with alternating colors
    for row in range(5, len(col_def_df) + 5):
        for col in range(1, 4):
            cell = col_def_worksheet.cell(row=row, column=col)
            
            if col == 1:  # Table column
                cell.fill = PatternFill(start_color="E1BEE7", end_color="E1BEE7", fill_type="solid")  # Light purple
                cell.font = Font(bold=True, size=10)
                cell.alignment = Alignment(horizontal="center")
            elif col == 2:  # Column Name column
                cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")  # Very light purple
                cell.font = Font(bold=True, size=10)
            else:  # Description column
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.font = Font(size=10)
    
    # Adjust column widths
    col_def_worksheet.column_dimensions['A'].width = 8   # Table column
    col_def_worksheet.column_dimensions['B'].width = 30  # Column Name column
    col_def_worksheet.column_dimensions['C'].width = 100 # Description column
    
    # Set row heights for better readability
    for row in range(5, len(col_def_df) + 5):
        col_def_worksheet.row_dimensions[row].height = 40

def create_toc_worksheet(writer):
    """Create Table of Contents worksheet"""
    toc_data = {
        'Table': ['S1', 'S2', 'S3', 'S4', 'S5', 'S6', 'S7', 'S8', 'S9', 'S10', 'S11', 'S12', 'S13', 'S14', 'S15', 'S16', 'S17', 'S18',
                  '', 'NOTE'],
        'Title': [
            'Brain Microproteins Discovery Summary',
            'Proteomics Results Summary', 
            'Tryptic Peptide Microproteins',
            'RP3 Results Summary',
            'ShortStop Microproteins Summary',
            'Single-cell RNA Enrichment Summary',
            'Short-Read RNA-seq Transcriptomics Results',
            'TSS Motif Prevalence',
            'RBP Splice-Site Motifs',
            'Co-expression Analysis AD (Target ENST00000347364)',
            'Co-expression Analysis Non-AD (Target ENST00000347364)',
            'Co-expression Analysis AD (Target ENSG00000125863)',
            'Co-expression Analysis Combined (Target ENSG00000125863)',
            'Outer Mitochondrial Membrane Differential Proteins',
            'Long-Read RNA-seq Transcriptomics Results',
            'RP3 Pseudogene-derived ORFs',
            'TMT ROSMAP Donor Metadata',
            'smORF Type Definitions',
            '',
            'Status Column Legend'
        ],
        'Description': [
            'Comprehensive annotation and classification of microproteins identified across all analyses',
            'Mass spectrometry evidence for microprotein expression with peptide-spectrum matches',
            'Tryptic peptide analysis of microproteins with unique peptide sequences',
            'Ribosome footprint quantification evidence for microprotein translation with proteomics validation',
            'Machine learning classification and annotation of smORFs with evidence of translation through RiboCode',
            'Cell-type specific expression patterns of microproteins in brain tissue',
            'Differential expression analysis of microproteins in Alzheimer\'s Disease vs controls',
            'Promoter motif prevalence at predicted smORF transcription start sites',
            'Per-junction RNA-binding protein motif enrichment at smORF-spanning splice sites',
            'Gene co-expression networks in Alzheimer\'s Disease for Micro-MKKS63',
            'Gene co-expression networks in control samples for Micro-MKKS63',
            'Gene co-expression networks in Alzheimer\'s Disease for reviewed MKKS-570aa',
            'Combined gene co-expression networks for reviewed MKKS-570aa',
            'Proteomics analysis of mitochondrial membrane-associated microproteins',
            'Long-read sequencing analysis of microprotein transcript expression',
            'Ribosome footprint quantification evidence for translation of pseudogene-derived microproteins',
            'Demographic characteristics of TMT ROSMAP study participants by biological sex and diagnostic group',
            'Comprehensive definitions and descriptions of small open reading frame (smORF) classification categories used throughout the analysis',
            '',
            'Reviewed = Swiss-Prot entries. Unreviewed = TrEMBL and Novel (Salk-generated database or TrEMBL).'
        ]
    }
    
    toc_df = pd.DataFrame(toc_data)
    toc_df.to_excel(writer, sheet_name='TOC', index=False)
    
    # Style the TOC
    toc_worksheet = writer.sheets['TOC']
    
    # Add title
    toc_worksheet.insert_rows(1)
    toc_worksheet.insert_rows(1)
    toc_worksheet.insert_rows(1)  # Three rows for title
    
    title_cell = toc_worksheet.cell(row=1, column=1)
    title_cell.value = "TABLE OF CONTENTS - Brain Microproteins Supplementary Data"
    
    # Style title
    title_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")  # Dark indigo
    title_font = Font(color="FFFFFF", bold=True, size=16)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    toc_worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    toc_worksheet.row_dimensions[1].height = 50
    
    # Style headers (row 4)
    header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")  # Indigo
    header_font = Font(color="FFFFFF", bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col in range(1, 4):  # 3 columns
        cell = toc_worksheet.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Style data rows
    for row in range(5, len(toc_df) + 5):
        for col in range(1, 4):
            cell = toc_worksheet.cell(row=row, column=col)
            if col == 1:  # Table column
                cell.fill = PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid")
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            elif col == 2:  # Title column
                cell.fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
                cell.font = Font(bold=True)
            else:  # Description column
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Adjust column widths
    toc_worksheet.column_dimensions['A'].width = 8
    toc_worksheet.column_dimensions['B'].width = 35
    toc_worksheet.column_dimensions['C'].width = 80
    
    # Set row heights for better readability
    for row in range(5, len(toc_df) + 5):
        toc_worksheet.row_dimensions[row].height = 60

def add_table_title(worksheet, table_id, num_columns):
    """Add title row at the top of the worksheet"""
    # Insert a new row at the top
    worksheet.insert_rows(1)
    worksheet.insert_rows(1)  # Insert two rows for title and spacing
    
    # Get the full title
    full_title = get_full_table_title(table_id)
    title_text = f"Supplementary Table {table_id}. {full_title}"
    
    # Add title to cell A1
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.value = title_text
    
    # Style the title
    title_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")  # Dark blue
    title_font = Font(color="FFFFFF", bold=True, size=14)
    title_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    title_cell.fill = title_fill
    title_cell.font = title_font
    title_cell.alignment = title_alignment
    
    # Merge title across all columns
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    
    # Set row height for title
    worksheet.row_dimensions[1].height = 40

def get_column_color_scheme():
    """Define color schemes for different column types"""
    return {
        # Identifiers - Blue tones
        'id': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),  # Light blue
        'gene': PatternFill(start_color="E1F5FE", end_color="E1F5FE", fill_type="solid"),  # Cyan
        'sequence': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # Light purple
        
        # Statistical measures - Green tones
        'pvalue': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),  # Light green
        'padj': PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid"),  # Pale green
        'fdr': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),  # Light green
        
        # Fold changes - Orange/Red tones
        'fold': PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),  # Light orange
        'logfc': PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid"),  # Pale orange
        'log2fc': PatternFill(start_color="FFECB3", end_color="FFECB3", fill_type="solid"),  # Pale orange
        
        # Counts/Quantitative - Yellow tones
        'count': PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid"),  # Light yellow
        'psm': PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),  # Pale yellow
        'spectral': PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),  # Pale yellow
        
        # Classifications - Purple tones
        'database': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # Light purple
        'classification': PatternFill(start_color="EDE7F6", end_color="EDE7F6", fill_type="solid"),  # Pale purple
        'type': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),  # Light purple
        
        # Coordinates/Locations - Pink tones
        'chr': PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),  # Light pink
        'coordinate': PatternFill(start_color="F8BBD9", end_color="F8BBD9", fill_type="solid"),  # Pink
        'ucsc': PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),  # Light pink
        
        # Default
        'default': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
    }

def get_column_fill(column_name):
    """Determine fill color for a column based on its name"""
    colors = get_column_color_scheme()
    column_lower = column_name.lower()
    
    # Check for specific patterns in column names
    if column_lower == 'status':
        return colors['classification']  # Special highlighting for Status column
    elif 'click_ucsc' in column_lower or 'ucsc' in column_lower:
        return colors['ucsc']  # Special color for UCSC links
    elif any(term in column_lower for term in ['gene', 'symbol']):
        return colors['gene']
    elif any(term in column_lower for term in ['sequence', 'seq']):
        return colors['sequence']
    elif any(term in column_lower for term in ['pvalue', 'p_value', 'p.value', 'pval']):
        return colors['pvalue']
    elif any(term in column_lower for term in ['padj', 'p_adj', 'fdr', 'qvalue', 'q_value']):
        return colors['padj']
    elif any(term in column_lower for term in ['fold', 'fc', 'logfc', 'log2fc', 'log_fc']):
        return colors['logfc']
    elif any(term in column_lower for term in ['count', 'psm', 'spectral']):
        return colors['count']
    elif any(term in column_lower for term in ['database', 'classification', 'class', 'type']):
        return colors['classification']
    elif any(term in column_lower for term in ['chr', 'coordinate', 'position']):
        return colors['chr']
    else:
        return colors['default']

def style_worksheet(worksheet, df):
    """Apply styling to the worksheet (accounting for title rows)"""
    # Header styling (now on row 3 due to title)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")  # Dark green
    header_font = Font(color="FFFFFF", bold=True, size=12)  # White, bold
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style headers (row 3, after title and spacing)
    for col_num, column_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Style data columns based on content type
    for col_num, column_name in enumerate(df.columns, 1):
        column_fill = get_column_fill(column_name)
        
        # Apply to all data rows (start from row 4, after title and header)
        for row_num in range(4, len(df) + 4):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.fill = column_fill
            cell.border = thin_border
            
            # Center align certain column types
            if any(term in column_name.lower() for term in ['chr', 'classification', 'database', 'type']):
                cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths (skip merged cells)
    for col_num in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = worksheet.cell(row=3, column=col_num).column_letter  # Use header row to get column letter
        
        # Check data rows only (skip title rows)
        for row_num in range(3, len(df) + 4):  # Start from header row
            cell = worksheet.cell(row=row_num, column=col_num)
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
                
        # Set reasonable bounds for column width
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    """Generate master supplementary Excel file with multiple tabs"""

    parser = argparse.ArgumentParser(description="Generate supplementary tables Excel file")
    parser.add_argument(
        "--include-scrna",
        action="store_true",
        default=False,
        help="Include the scRNA Enrichment table (S6). Omitted by default."
    )
    args = parser.parse_args()
    include_scrna = args.include_scrna
    if include_scrna:
        print("ℹ️  --include-scrna flag set: scRNA Enrichment table (S6) will be included.")
    else:
        print("ℹ️  scRNA Enrichment table (S6) excluded by default. Use --include-scrna to include it.")
    
    # Output configuration - use local directory first to avoid Box timeout issues
    local_output_dir = Path.home() / "Desktop"
    output_file = local_output_dir / "Supplemental_Tables.xlsx"
    
    # Final destination in Box
    final_output_dir = Path("/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/supplementary")
    final_output_file = final_output_dir / "Supplemental_Tables.xlsx"
    
    # Create output directories if they don't exist
    local_output_dir.mkdir(parents=True, exist_ok=True)
    final_output_dir.mkdir(parents=True, exist_ok=True)
    
    # Define supplementary tables mapping
    supplementary_tables = {
        "S1": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/Annotations/Brain_Microproteins_Discovery_summary.csv",
            "description": "Brain Microproteins Discovery Summary"
        },
        "S2": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/Proteomics/Proteomics_Results_summary.csv",
            "description": "Proteomics Results Summary"
        },
        "S3": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Code/data/cleaned_tryptic_peptides_under_151aa.csv",
            "description": "Tryptic Peptides Summary"
        },
        "S4": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/RP3/RP3_Results_summary.csv",
            "description": "RP3 Results Summary"
        },
        "S5": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/Annotations/ShortStop_Microproteins_summary.csv",
            "description": "ShortStop Microproteins Summary"
        },
        "S6": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/scRNA_Enrichment/scRNA_Enrichment_summary.csv",
            "description": "scRNA Enrichment Summary",
            "optional_scrna": True,
        },
        "S7": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/Transcriptomics/Short-Read_Transcriptomics_Results_summary.csv",
            "description": "Short-Read Transcriptomics Results"
        },
        "S8": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Code/data/tss_motif_prevalence.csv",
            "description": "TSS Motif Prevalence"
        },
        "S9": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Code/data/rbp_splice_per_junction.csv",
            "description": "RBP Splice-Site Motifs"
        },
        "S10": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brain_smorfs/shortstop/deseq_on_rosmap/coexpression_outputs/target_ENST00000347364_7_chr20_10420549_10420737_F_2_P_3_M/coexpression_AD.csv",
            "description": "Coexpression Analysis - AD (Target ENST00000347364)"
        },
        "S11": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brain_smorfs/shortstop/deseq_on_rosmap/coexpression_outputs/target_ENST00000347364_7_chr20_10420549_10420737_F_2_P_3_M/coexpression_nonAD.csv",
            "description": "Coexpression Analysis - Non-AD (Target ENST00000347364)"
        },
        "S12": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brain_smorfs/shortstop/deseq_on_rosmap/coexpression_outputs/target_ENSG00000125863_20/coexpression_AD.csv",
            "description": "Coexpression Analysis - AD (Target ENSG00000125863)"
        },
        "S13": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brain_smorfs/shortstop/deseq_on_rosmap/coexpression_outputs/target_ENSG00000125863_20/coexpression_nonAD.csv",
            "description": "Coexpression Analysis - Non-AD (Target ENSG00000125863)"
        },
        "S14": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/apex_omm/omm_differential_proteins.csv",
            "description": "OMM Differential Proteins"
        },
        "S15": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/Transcriptomics/Long-Read_Transcriptomics_Results_summary.csv",
            "description": "Long-Read Transcriptomics Results"
        },
        "S16": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/RP3/RP3_psORFs.csv",
            "description": "RP3 psORFs"
        },
        "S17": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/supplementary/S14_TMT_ROSMAP_Donor_Metadata.csv",
            "description": "TMT ROSMAP Donor Metadata"
        },
        "S18": {
            "path": "/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Results/Annotations/smORF_type_definitions.csv",
            "description": "smORF Type Definitions"
        }
    }

    # Remove optional scRNA table unless explicitly requested
    if not include_scrna:
        supplementary_tables = {
            k: v for k, v in supplementary_tables.items()
            if not v.get("optional_scrna", False)
        }

    print("=" * 80)
    print("Generating Master Supplementary Tables Excel File")
    print("=" * 80)
    print(f"Output file: {output_file}")
    print(f"Total tables to process: {len(supplementary_tables)}")
    print()
    
    # Create Excel writer object
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        # Create Table of Contents first
        print("Creating Table of Contents...")
        create_toc_worksheet(writer)
        print("  ✅ SUCCESS: Table of Contents created")
        print()
        
        # Create Column Definitions worksheet
        print("Creating Column Definitions...")
        create_column_definitions_worksheet(writer)
        print("  ✅ SUCCESS: Column Definitions created")
        print()
        
        # Track successful and failed loads
        successful_tables = []
        failed_tables = []
        
        # Process each supplementary table
        for table_id, table_info in supplementary_tables.items():
            file_path = Path(table_info["path"])
            description = table_info["description"]
            
            print(f"Processing {table_id}: {description}")
            print(f"  Source: {file_path.name}")
            
            try:
                # Check if file exists
                if not file_path.exists():
                    print(f"  ❌ ERROR: File not found - {file_path}")
                    failed_tables.append((table_id, "File not found", str(file_path)))
                    continue
                
                # Load CSV data
                df = pd.read_csv(file_path, low_memory=False)
                
                # Remove any unnamed columns (row indices)
                unnamed_cols = [col for col in df.columns if col.startswith('Unnamed')]
                if unnamed_cols:
                    df = df.drop(columns=unnamed_cols)
                    print(f"  🧹 Removed {len(unnamed_cols)} unnamed columns")
                
                # Move CLICK_UCSC column to front if it exists
                df = move_click_ucsc_to_front(df)
                
                # Add Status column as first column (after CLICK_UCSC if present)
                df = add_status_column(df)
                
                # Apply table-specific filtering (S8 nanopore filter)
                df = filter_s8_nanopore_data(df, table_id)
                
                # Enrich S3 (Tryptic Peptides) with PROSIT quality metrics from master
                if table_id == "S3":
                    try:
                        master_path = Path("/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Code/data/microprotein_master.csv")
                        prosit_cols = ['sequence', 'SA_degrees', 'SA_normalized', 'Match_coverage', 'Match_coverage_pct', 'Confidence']
                        master_prosit = pd.read_csv(master_path, low_memory=False, usecols=prosit_cols)
                        master_prosit = master_prosit.drop_duplicates(subset='sequence')
                        before_cols = len(df.columns)
                        df = df.merge(master_prosit, on='sequence', how='left')
                        print(f"  🔬 Merged PROSIT quality columns from master ({len(df.columns) - before_cols} columns added)")
                    except Exception as e:
                        print(f"  ⚠️  Warning: Could not merge PROSIT columns - {str(e)}")

                # Enrich S8/S9 with Status, CLICK_UCSC, gene_symbol from master
                if table_id in ("S8", "S9"):
                    try:
                        master_path = Path("/Users/brendanmiller/Library/CloudStorage/Box-Box/brendan_alan_shared_folder/AD_paper/Github/Code/data/microprotein_master.csv")
                        annot_cols = ['sequence', 'CLICK_UCSC', 'gene_symbol', 'Database']
                        master_annot = pd.read_csv(master_path, low_memory=False, usecols=annot_cols)
                        master_annot = master_annot.drop_duplicates(subset='sequence')
                        df = df.merge(master_annot, on='sequence', how='left')
                        # Add Status from merged Database column, then move key cols to front
                        df = add_status_column(df)
                        df = move_click_ucsc_to_front(df)
                        # Ensure gene_symbol appears right after sequence
                        cols = list(df.columns)
                        for move_col in ['gene_symbol', 'sequence']:
                            if move_col in cols:
                                cols.insert(0, cols.pop(cols.index(move_col)))
                        df = df[cols]
                        # Drop raw Database column (Status already captures it)
                        if 'Database' in df.columns:
                            df = df.drop(columns=['Database'])
                        print(f"  🔗 Merged Status, CLICK_UCSC, gene_symbol from master")
                    except Exception as e:
                        print(f"  ⚠️  Warning: Could not merge annotation columns - {str(e)}")

                # Apply table-specific sorting
                df = apply_table_sorting(df, table_id)
                
                # Write to Excel sheet with descriptive name and styling
                # Create sheet name with brief description (Excel has 31 char limit)
                brief_desc = get_brief_description(table_id, description)
                sheet_name = f"{table_id} {brief_desc}"
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet and apply formatting
                worksheet = writer.sheets[sheet_name]
                
                # Add title row at the top
                add_table_title(worksheet, table_id, len(df.columns))
                
                # Apply styling to the worksheet
                style_worksheet(worksheet, df)
                
                print(f"  ✅ SUCCESS: {len(df):,} rows, {len(df.columns)} columns")
                successful_tables.append((table_id, description, len(df), len(df.columns)))
                
            except Exception as e:
                print(f"  ❌ ERROR: {str(e)}")
                failed_tables.append((table_id, str(e), str(file_path)))
            
            print()
    
    # Generate summary report
    print("=" * 80)
    print("SUMMARY REPORT")
    print("=" * 80)
    print(f"Successfully processed: {len(successful_tables)}/{len(supplementary_tables)} tables")
    print(f"Output file created: {output_file}")
    print()
    
    if successful_tables:
        print("✅ SUCCESSFUL TABLES:")
        total_rows = 0
        for table_id, desc, rows, cols in successful_tables:
            print(f"  {table_id}: {desc} ({rows:,} rows, {cols} cols)")
            total_rows += rows
        print(f"  TOTAL ROWS ACROSS ALL TABLES: {total_rows:,}")
        print()
    
    if failed_tables:
        print("❌ FAILED TABLES:")
        for table_id, error, path in failed_tables:
            print(f"  {table_id}: {error}")
            print(f"    Path: {path}")
        print()
        print("🛠️  To fix failed tables:")
        print("   1. Ensure all analysis scripts have been run successfully")
        print("   2. Check file paths are correct")
        print("   3. Verify CSV files are properly formatted")
        print()
    
    # File size info
    if output_file.exists():
        file_size_mb = output_file.stat().st_size / (1024 * 1024)
        print(f"📁 Output file size: {file_size_mb:.2f} MB")
    
    print("=" * 80)
    print("Excel generation completed!")
    print("=" * 80)
    
    # Copy file to final destination in Box
    try:
        print(f"Copying file to final destination: {final_output_file}")
        shutil.copy2(output_file, final_output_file)
        print("✅ File successfully copied to Box storage")
        print(f"📁 Final location: {final_output_file}")
    except Exception as e:
        print(f"⚠️  Warning: Could not copy to Box storage: {e}")
        print(f"📁 File saved locally at: {output_file}")
    
    return len(failed_tables) == 0  # Return True if all successful

if __name__ == "__main__":
    success = main()
    if not success:
        sys.exit(1)  # Exit with error code if any tables failed